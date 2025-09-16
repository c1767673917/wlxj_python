from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify
from flask_login import login_required, current_user
from models import db, Order, Supplier, order_suppliers
from utils.auth import business_type_filter
from datetime import datetime, date
from sqlalchemy import or_, func
import requests
import json
import logging
import traceback
import time
from sqlalchemy.exc import SQLAlchemyError, IntegrityError
from requests.exceptions import RequestException, Timeout, ConnectionError
from utils.error_codes import ErrorCode, ErrorHandler, ErrorResponseHelper

# 创建蓝图
order_bp = Blueprint('order', __name__, url_prefix='/orders')

@order_bp.route('/')
@login_required
def index():
    """订单列表页面 - 支持日期和关键词筛选，增强错误处理"""
    try:
        # 安全地获取参数，设置合理的默认值
        page = request.args.get('page', 1, type=int)
        if page < 1:
            page = 1
        elif page > 1000:  # 防止恶意的过大页码
            page = 1000
            
        status = request.args.get('status', '').strip()
        start_date = request.args.get('start_date', '').strip()
        end_date = request.args.get('end_date', '').strip()
        keyword = request.args.get('keyword', '').strip()
        date_quick = request.args.get('date_quick', '').strip()
        
        # 验证状态参数
        valid_statuses = ['', 'active', 'completed', 'cancelled']
        if status not in valid_statuses:
            logging.warning(f"用户提供了无效的状态参数: {status}")
            status = ''
        
        # 限制关键词长度，防止过长的搜索字符串
        if keyword and len(keyword) > 100:
            logging.warning(f"用户提供了过长的关键词: {len(keyword)}字符")
            keyword = keyword[:100]
            flash('搜索关键词过长，已自动截取前100个字符', 'warning')
        
        # 处理快捷日期选项
        if date_quick:
            quick_start, quick_end = process_quick_date(date_quick)
            if quick_start and quick_end:
                start_date, end_date = quick_start, quick_end
                logging.debug(f"应用快捷日期选项: {date_quick} -> {start_date} to {end_date}")
            else:
                logging.warning(f"快捷日期选项处理失败: {date_quick}")
                flash('快捷日期设置失败，请手动选择日期', 'warning')
        
        # 构建查询
        query = Order.query
        query = business_type_filter(query, Order)
        
        # 状态筛选
        if status:
            query = query.filter_by(status=status)
            logging.debug(f"应用状态筛选: {status}")
        
        # 日期范围筛选 - 使用增强的错误处理
        try:
            query = apply_date_filter(query, start_date, end_date)
        except Exception as e:
            logging.error(f"日期筛选处理失败: {str(e)}")
            flash('日期筛选处理失败，显示所有订单', 'error')
            # 重置日期参数，避免继续错误
            start_date = end_date = ''
        
        # 关键词搜索 - 使用增强的错误处理
        try:
            query = apply_keyword_search(query, keyword)
        except Exception as e:
            logging.error(f"关键词搜索处理失败: {str(e)}")
            flash('搜索功能暂时不可用，显示所有订单', 'error')
            keyword = ''
        
        # 执行分页查询，增加异常处理和性能监控
        try:
            # 查询性能监控
            start_time = time.time()
            
            orders = query.order_by(Order.created_at.desc()).paginate(
                page=page, per_page=10, error_out=False)
            
            # 性能监控和日志记录
            query_time = time.time() - start_time
            if query_time > 1.0:  # 查询时间超过1秒记录警告
                logging.warning(f"慢查询检测: 订单列表查询耗时 {query_time:.2f}秒")
            logging.debug(f"订单列表查询耗时: {query_time:.3f}秒")
            
            # 检查分页结果是否有效
            if orders.total == 0 and (status or start_date or end_date or keyword):
                logging.info(f"筛选条件未找到结果 - 状态:{status}, 日期:{start_date}-{end_date}, 关键词:{keyword}")
            elif orders.total > 0:
                logging.debug(f"查询完成，找到{orders.total}条记录，显示第{page}页")
                
        except Exception as e:
            logging.error(f"分页查询执行失败: {str(e)}")
            flash('查询失败，请稍后重试', 'error')
            # 回退到简单查询
            try:
                orders = Order.query.filter(
                    business_type_filter(Order.query, Order).statement.whereclause
                ).order_by(Order.created_at.desc()).paginate(
                    page=1, per_page=10, error_out=False)
            except Exception as fallback_error:
                logging.error(f"回退查询也失败: {str(fallback_error)}")
                # 创建空的分页对象
                from flask_sqlalchemy import Pagination
                orders = type('MockPagination', (), {
                    'items': [], 'total': 0, 'pages': 0, 
                    'has_prev': False, 'has_next': False,
                    'prev_num': None, 'next_num': None,
                    'page': 1, 'per_page': 10,
                    'iter_pages': lambda: []
                })()
        
        return render_template('orders/index.html', 
                             orders=orders, 
                             status=status,
                             start_date=start_date,
                             end_date=end_date,
                             keyword=keyword,
                             date_quick=date_quick)
    
    except Exception as e:
        logging.error(f"订单列表页面发生未知错误: {str(e)}")
        logging.error(f"错误详情: {traceback.format_exc()}")
        flash('页面加载失败，请刷新页面重试', 'error')
        
        # 返回最基本的页面
        try:
            basic_orders = Order.query.order_by(Order.created_at.desc()).limit(10).all()
            # 构造基础分页对象
            orders = type('BasicPagination', (), {
                'items': basic_orders, 'total': len(basic_orders), 'pages': 1,
                'has_prev': False, 'has_next': False,
                'prev_num': None, 'next_num': None,
                'page': 1, 'per_page': 10,
                'iter_pages': lambda: [1]
            })()
            
            return render_template('orders/index.html', 
                                 orders=orders, 
                                 status='',
                                 start_date='',
                                 end_date='',
                                 keyword='',
                                 date_quick='')
        except Exception as final_error:
            logging.error(f"基础回退也失败: {str(final_error)}")
            return "系统错误，请联系管理员", 500

@order_bp.route('/new', methods=['GET', 'POST'])
@login_required
def create():
    """创建新订单 - 带完整异常处理和事务回滚"""
    if request.method == 'POST':
        # 开始事务
        try:
            warehouse = request.form.get('warehouse', '').strip()
            goods = request.form.get('goods', '').strip()
            delivery_address = request.form.get('delivery_address', '').strip()
            supplier_ids = request.form.getlist('supplier_ids')
            
            # 管理员可以选择业务类型，普通用户使用自己的业务类型
            if current_user.is_admin():
                business_type = request.form.get('business_type', 'oil')
                if business_type not in ['oil', 'fast_moving']:
                    flash('无效的业务类型', 'error')
                    query = Supplier.query
                    suppliers = business_type_filter(query, Supplier).all()
                    return render_template('orders/create.html', suppliers=suppliers)
            else:
                business_type = current_user.business_type
            
            # 数据验证 - 使用统一错误码
            if not warehouse:
                ErrorResponseHelper.flash_error_message(ErrorCode.VAL_001, "仓库信息")
                query = Supplier.query.filter_by(business_type=business_type)
                suppliers = query.all() if current_user.is_admin() else business_type_filter(Supplier.query, Supplier).all()
                return render_template('orders/create.html', suppliers=suppliers)
                
            if not goods:
                ErrorResponseHelper.flash_error_message(ErrorCode.VAL_001, "货物信息")
                query = Supplier.query.filter_by(business_type=business_type)
                suppliers = query.all() if current_user.is_admin() else business_type_filter(Supplier.query, Supplier).all()
                return render_template('orders/create.html', suppliers=suppliers)
                
            if not delivery_address:
                ErrorResponseHelper.flash_error_message(ErrorCode.VAL_001, "收货地址")
                query = Supplier.query.filter_by(business_type=business_type)
                suppliers = query.all() if current_user.is_admin() else business_type_filter(Supplier.query, Supplier).all()
                return render_template('orders/create.html', suppliers=suppliers)
            
            if not supplier_ids:
                ErrorResponseHelper.flash_error_message(ErrorCode.VAL_001, "请至少选择一个供应商")
                # 根据业务类型获取供应商
                query = Supplier.query.filter_by(business_type=business_type)
                suppliers = query.all() if current_user.is_admin() else business_type_filter(Supplier.query, Supplier).all()
                return render_template('orders/create.html', suppliers=suppliers)
            
            # 验证供应商ID是否有效
            supplier_ids = [int(sid) for sid in supplier_ids if sid.isdigit()]
            if not supplier_ids:
                ErrorResponseHelper.flash_error_message(ErrorCode.VAL_008, "供应商ID格式无效")
                query = Supplier.query.filter_by(business_type=business_type)
                suppliers = query.all() if current_user.is_admin() else business_type_filter(Supplier.query, Supplier).all()
                return render_template('orders/create.html', suppliers=suppliers)
            
            # 创建订单对象
            order = Order(
                order_no=Order.generate_temp_order_no(),
                warehouse=warehouse,
                goods=goods,
                delivery_address=delivery_address,
                user_id=current_user.id,
                business_type=business_type
            )
            
            # 数据验证
            validation_errors = order.validate_order_data()
            if validation_errors:
                for error in validation_errors:
                    flash(error, 'error')
                query = Supplier.query.filter_by(business_type=business_type)
                suppliers = query.all() if current_user.is_admin() else business_type_filter(Supplier.query, Supplier).all()
                return render_template('orders/create.html', suppliers=suppliers)
            
            # 验证供应商是否属于指定业务类型
            selected_suppliers = Supplier.query.filter(Supplier.id.in_(supplier_ids), Supplier.business_type == business_type).all()
            
            if len(selected_suppliers) != len(supplier_ids):
                flash('选择的供应商中包含无效项目', 'error')
                query = Supplier.query.filter_by(business_type=business_type)
                suppliers = query.all() if current_user.is_admin() else business_type_filter(Supplier.query, Supplier).all()
                return render_template('orders/create.html', suppliers=suppliers)
            
            # 开始数据库事务
            db.session.add(order)
            db.session.flush()  # 获取订单ID但不提交
            
            # 生成正式订单号（基于ID确保唯一性）
            order.order_no = order.generate_order_no()
            
            # 关联选中的供应商
            for supplier in selected_suppliers:
                order.suppliers.append(supplier)
            
            # 提交事务
            db.session.commit()
            
            logging.info(f"订单创建成功: {order.order_no}, 用户: {current_user.id}, 供应商数量: {len(selected_suppliers)}")
            
            # 异步发送通知给供应商（不影响主流程）
            try:
                success_count, failed_suppliers = notify_suppliers(order, selected_suppliers)
                if failed_suppliers:
                    notification_status = f"已通知 {success_count} 个供应商，{len(failed_suppliers)} 个通知失败"
                else:
                    notification_status = f"已成功通知 {success_count} 个供应商"
            except Exception as notify_error:
                logging.error(f"发送供应商通知失败: {str(notify_error)}")
                notification_status = "订单创建成功，但通知发送失败"
            
            flash(f'订单 {order.order_no} 创建成功，{notification_status}', 'success')
            return redirect(url_for('order.detail', order_id=order.id))
            
        except IntegrityError as e:
            db.session.rollback()
            error_response, _ = ErrorHandler.handle_database_error(e)
            ErrorResponseHelper.flash_error_message((error_response["error_code"], error_response["error_message"]))
        except SQLAlchemyError as e:
            db.session.rollback()
            ErrorResponseHelper.flash_error_message(ErrorCode.SYS_002, "数据库操作失败，请稍后重试")
        except ValueError as e:
            db.session.rollback()
            ErrorResponseHelper.flash_error_message(ErrorCode.VAL_003, str(e))
        except Exception as e:
            db.session.rollback()
            logging.error(f"订单创建失败 - 未知错误: {str(e)}")
            logging.error(f"错误详情: {traceback.format_exc()}")
            ErrorResponseHelper.flash_error_message(ErrorCode.SYS_005, "系统异常，请联系管理员")
        
        # 出错时返回表单
        try:
            if current_user.is_admin():
                suppliers = Supplier.query.all()
            else:
                query = Supplier.query
                suppliers = business_type_filter(query, Supplier).all()
        except Exception:
            suppliers = []
        return render_template('orders/create.html', suppliers=suppliers)
    
    # GET请求 - 显示创建表单
    try:
        if current_user.is_admin():
            suppliers = Supplier.query.all()
        else:
            query = Supplier.query
            suppliers = business_type_filter(query, Supplier).all()
    except SQLAlchemyError as e:
        logging.error(f"获取供应商列表失败: {str(e)}")
        flash('获取供应商列表失败，请稍后重试', 'error')
        suppliers = []
    
    return render_template('orders/create.html', suppliers=suppliers)

@order_bp.route('/<int:order_id>')
@login_required
def detail(order_id):
    """订单详情页面 - 使用缓存机制优化性能"""
    try:
        query = Order.query.filter_by(id=order_id)
        order = business_type_filter(query, Order).first_or_404()
        
        # 使用缓存机制获取Quote模型，提升性能
        Quote = Order._get_quote_model()
        quotes = Quote.query.filter_by(order_id=order.id).order_by(Quote.price.asc()).all()
        
        # 记录性能统计信息
        cache_stats = Order.get_cache_stats()
        logging.debug(f"订单详情页面加载完成，缓存命中率: {cache_stats['hit_rate_percent']}%")
        
        return render_template('orders/detail.html', order=order, quotes=quotes)
        
    except Exception as e:
        logging.error(f"加载订单详情失败 (订单ID: {order_id}): {str(e)}")
        flash('加载订单详情失败，请稍后重试', 'error')
        return redirect(url_for('order.index'))

@order_bp.route('/<int:order_id>/edit', methods=['GET', 'POST'])
@login_required
def edit(order_id):
    """编辑订单 - 带异常处理"""
    try:
        query = Order.query.filter_by(id=order_id)
        order = business_type_filter(query, Order).first_or_404()
        
        if order.status != 'active':
            flash('只能编辑活跃状态的订单', 'error')
            return redirect(url_for('order.detail', order_id=order.id))
        
        if request.method == 'POST':
            try:
                warehouse = request.form.get('warehouse', '').strip()
                goods = request.form.get('goods', '').strip()
                delivery_address = request.form.get('delivery_address', '').strip()
                
                # 数据验证
                if not all([warehouse, goods, delivery_address]):
                    flash('请填写所有必填字段', 'error')
                    return render_template('orders/edit.html', order=order)
                
                # 长度验证
                if len(warehouse) > 200:
                    flash('仓库信息长度不能超过200字符', 'error')
                    return render_template('orders/edit.html', order=order)
                
                if len(delivery_address) > 300:
                    flash('收货地址长度不能超过300字符', 'error')
                    return render_template('orders/edit.html', order=order)
                
                # 保存原始数据用于回滚
                original_warehouse = order.warehouse
                original_goods = order.goods
                original_delivery_address = order.delivery_address
                
                # 更新订单数据
                order.warehouse = warehouse
                order.goods = goods
                order.delivery_address = delivery_address
                
                # 提交事务
                db.session.commit()
                
                logging.info(f"订单编辑成功: {order.order_no}, 用户: {current_user.id}")
                flash('订单信息更新成功', 'success')
                return redirect(url_for('order.detail', order_id=order.id))
                
            except SQLAlchemyError as e:
                db.session.rollback()
                logging.error(f"订单编辑失败 - 数据库错误: {str(e)}")
                flash('订单更新失败：数据库错误，请稍后重试', 'error')
                return render_template('orders/edit.html', order=order)
            except Exception as e:
                db.session.rollback()
                logging.error(f"订单编辑失败 - 未知错误: {str(e)}")
                flash('订单更新失败：系统错误，请联系管理员', 'error')
                return render_template('orders/edit.html', order=order)
        
        return render_template('orders/edit.html', order=order)
        
    except Exception as e:
        logging.error(f"获取订单详情失败: {str(e)}")
        flash('获取订单信息失败', 'error')
        return redirect(url_for('order.index'))

@order_bp.route('/<int:order_id>/select-supplier', methods=['POST'])
@login_required
def select_supplier(order_id):
    """选择中标供应商 - 使用缓存机制优化性能"""
    try:
        query = Order.query.filter_by(id=order_id)
        order = business_type_filter(query, Order).first_or_404()
        
        supplier_id = request.form.get('supplier_id', type=int)
        price = request.form.get('price', type=float)
        
        if not supplier_id or not price:
            flash('请选择供应商和确认价格', 'error')
            return redirect(url_for('order.detail', order_id=order.id))
        
        # 验证供应商是否有该订单的报价
        # 使用缓存机制获取Quote模型，提升性能
        Quote = Order._get_quote_model()
        quote = Quote.query.filter_by(order_id=order.id, supplier_id=supplier_id).first()
        if not quote:
            flash('所选供应商没有该订单的报价', 'error')
            return redirect(url_for('order.detail', order_id=order.id))
        
        # 验证价格有效性（允许价格协商，不要求严格匹配）
        if price <= 0:
            flash('确认价格必须大于0', 'error')
            return redirect(url_for('order.detail', order_id=order.id))
        
        # 如果确认价格与报价不一致，记录日志用于审计
        original_price = float(quote.price)
        if abs(original_price - price) > 0.01:
            logging.info(f"订单 {order.order_no} 价格协商：原报价 {original_price}，确认价格 {price}，供应商ID: {supplier_id}")
            # 在flash消息中提示价格变更
            flash_message_suffix = f"（原报价：{original_price:.2f}元，确认价格：{price:.2f}元）"
        else:
            flash_message_suffix = ""
        
        # 更新订单状态
        order.selected_supplier_id = supplier_id
        order.selected_price = price
        order.status = 'completed'
        
        db.session.commit()
        
        # 记录性能统计
        cache_stats = Order.get_cache_stats()
        logging.info(f"订单 {order.order_no} 已完成，选择供应商ID: {supplier_id}，缓存命中率: {cache_stats['hit_rate_percent']}%")
        
        flash(f'已选择中标供应商，订单已完成{flash_message_suffix}', 'success')
        return redirect(url_for('order.detail', order_id=order.id))
        
    except Exception as e:
        db.session.rollback()
        logging.error(f"选择中标供应商失败 (订单ID: {order_id}): {str(e)}")
        flash('操作失败，请稍后重试', 'error')
        return redirect(url_for('order.detail', order_id=order_id))

@order_bp.route('/<int:order_id>/cancel', methods=['POST'])
@login_required
def cancel(order_id):
    """删除订单 - 物理删除订单及相关数据"""
    try:
        # 获取订单并验证权限
        query = Order.query.filter_by(id=order_id)
        order = business_type_filter(query, Order).first_or_404()
        
        # 验证订单状态 - 已完成订单不能删除
        if order.status == 'completed':
            flash('已完成的订单无法删除', 'error')
            return redirect(url_for('order.detail', order_id=order.id))
        
        # 记录删除前的信息用于日志
        order_no = order.order_no
        order_goods = order.goods[:50]  # 截取前50个字符
        quote_count = order.get_quote_count()
        order_status = order.status
        
        # 物理删除订单（级联删除Quote记录）
        db.session.delete(order)
        db.session.commit()
        
        # 记录删除操作日志
        logging.info(f"订单删除成功: {order_no}, 状态: {order_status}, 货物: {order_goods}, 报价数: {quote_count}, 操作用户: {current_user.id}")
        
        flash(f'订单 {order_no} 已成功删除', 'success')
        return redirect(url_for('order.index'))
        
    except Exception as e:
        db.session.rollback()
        logging.error(f"删除订单失败 (订单ID: {order_id}): {str(e)}")
        flash('删除订单失败，请稍后重试', 'error')
        return redirect(url_for('order.detail', order_id=order_id))

@order_bp.route('/<int:order_id>/add-suppliers', methods=['GET', 'POST'])
@login_required
def add_suppliers(order_id):
    """为订单添加更多供应商"""
    query = Order.query.filter_by(id=order_id)
    order = business_type_filter(query, Order).first_or_404()
    
    if order.status != 'active':
        flash('只能为活跃状态的订单添加供应商', 'error')
        return redirect(url_for('order.detail', order_id=order.id))
    
    if request.method == 'POST':
        supplier_ids = request.form.getlist('supplier_ids')
        
        if not supplier_ids:
            flash('请至少选择一个供应商', 'error')
            return redirect(url_for('order.add_suppliers', order_id=order.id))
        
        # 获取新的供应商（排除已关联的）
        current_supplier_ids = [s.id for s in order.suppliers]
        new_supplier_ids = [int(sid) for sid in supplier_ids if int(sid) not in current_supplier_ids]
        
        if not new_supplier_ids:
            flash('所选供应商已经关联到此订单', 'warning')
            return redirect(url_for('order.detail', order_id=order.id))
        
        # 添加新供应商
        new_suppliers = Supplier.query.filter(Supplier.id.in_(new_supplier_ids)).all()
        for supplier in new_suppliers:
            order.suppliers.append(supplier)
        
        db.session.commit()
        
        # 通知新供应商
        try:
            success_count, failed_suppliers = notify_suppliers(order, new_suppliers)
            if failed_suppliers:
                flash(f'已添加 {len(new_suppliers)} 个供应商，通知发送成功 {success_count} 个，失败 {len(failed_suppliers)} 个', 'warning')
            else:
                flash(f'已添加 {len(new_suppliers)} 个供应商，并成功发送通知', 'success')
        except Exception as e:
            logging.error(f"发送供应商通知异常: {str(e)}")
            flash(f'已添加 {len(new_suppliers)} 个供应商，但通知发送失败', 'warning')
        return redirect(url_for('order.detail', order_id=order.id))
    
    # 获取可添加的供应商（排除已关联的）
    current_supplier_ids = [s.id for s in order.suppliers]
    query = Supplier.query.filter(~Supplier.id.in_(current_supplier_ids))
    available_suppliers = business_type_filter(query, Supplier).all()
    
    return render_template('orders/add_suppliers.html', order=order, suppliers=available_suppliers)

@order_bp.route('/<int:order_id>/reset-selection', methods=['POST'])
@login_required
def reset_selection(order_id):
    """取消选择的供应商，重新激活订单"""
    try:
        query = Order.query.filter_by(id=order_id)
        order = business_type_filter(query, Order).first_or_404()
        
        # 验证订单状态
        if order.status != 'completed':
            flash('只能取消已完成订单的供应商选择', 'error')
            return redirect(url_for('order.detail', order_id=order.id))
        
        if not order.selected_supplier_id:
            flash('订单没有选择供应商，无法取消选择', 'error')
            return redirect(url_for('order.detail', order_id=order.id))
        
        # 记录取消前的信息
        old_supplier_name = order.selected_supplier.name if order.selected_supplier else '未知'
        old_price = order.selected_price
        
        # 重新激活订单
        order.reset_to_active()
        db.session.commit()
        
        logging.info(f"管理员 {current_user.username} 取消了订单 {order.order_no} 的供应商选择 - 原选择: {old_supplier_name}, 价格: {old_price}")
        flash(f'已取消供应商选择，订单重新激活为进行中状态', 'success')
        return redirect(url_for('order.detail', order_id=order.id))
        
    except ValueError as e:
        flash(str(e), 'error')
        return redirect(url_for('order.detail', order_id=order_id))
    except Exception as e:
        db.session.rollback()
        logging.error(f"取消供应商选择失败 (订单ID: {order_id}): {str(e)}")
        flash('操作失败，请稍后重试', 'error')
        return redirect(url_for('order.detail', order_id=order_id))

def process_quick_date(date_quick):
    """处理快捷日期选项 - 增强版本，支持错误处理和边缘情况"""
    try:
        today = date.today()
        
        # 验证系统日期是否正常
        if not today or today.year < 2020 or today.year > 2050:
            logging.error(f"系统日期异常: {today}")
            return '', ''
        
        if date_quick == 'today':
            date_str = today.strftime('%Y-%m-%d')
            return date_str, date_str
        elif date_quick == 'this_month':
            try:
                # 处理本月起始日期，考虑各种边缘情况
                start = today.replace(day=1)
                
                # 验证生成的日期是否有效
                if start > today:
                    logging.error(f"月初日期大于当前日期: start={start}, today={today}")
                    return '', ''
                
                start_str = start.strftime('%Y-%m-%d')
                end_str = today.strftime('%Y-%m-%d')
                
                # 二次验证字符串格式
                import re
                date_pattern = re.compile(r'^\d{4}-\d{2}-\d{2}$')
                if not date_pattern.match(start_str) or not date_pattern.match(end_str):
                    logging.error(f"日期格式异常: start={start_str}, end={end_str}")
                    return '', ''
                
                return start_str, end_str
            except ValueError as e:
                logging.error(f"创建月初日期失败: {str(e)}")
                return '', ''
        else:
            logging.warning(f"不支持的快捷日期选项: {date_quick}")
            return '', ''
            
    except Exception as e:
        logging.error(f"处理快捷日期选项时发生错误: {str(e)}")
        return '', ''

def apply_date_filter(query, start_date, end_date):
    """应用日期范围筛选 - 增强版本，支持全面的验证和错误处理"""
    import re
    from datetime import timedelta
    
    start_dt = None
    end_dt = None
    
    # 日期格式验证正则表达式
    date_pattern = re.compile(r'^\d{4}-\d{2}-\d{2}$')
    
    # 处理开始日期
    if start_date:
        start_date = start_date.strip()
        
        if not start_date:
            # 空字符串，跳过处理
            pass
        elif not date_pattern.match(start_date):
            logging.warning(f"开始日期格式无效: {start_date}")
            flash('开始日期格式无效，请使用YYYY-MM-DD格式', 'error')
            return query
        else:
            try:
                start_dt = datetime.strptime(start_date, '%Y-%m-%d')
                
                # 验证日期合理性
                current_year = datetime.now().year
                if start_dt.year < 2020 or start_dt.year > current_year + 1:
                    logging.warning(f"开始日期超出合理范围: {start_date}")
                    flash('开始日期超出有效范围，请选择2020年到明年之间的日期', 'error')
                    return query
                
                # 检查是否是未来日期
                if start_dt.date() > date.today():
                    logging.info(f"用户选择了未来的开始日期: {start_date}")
                    flash('开始日期不能是未来日期', 'warning')
                
                query = query.filter(Order.created_at >= start_dt)
                logging.debug(f"应用开始日期筛选: {start_date}")
                
            except ValueError as e:
                logging.error(f"解析开始日期失败: {start_date}, 错误: {str(e)}")
                flash('开始日期无效，请检查日期是否存在（如2月30日）', 'error')
                return query
            except Exception as e:
                logging.error(f"处理开始日期时发生未知错误: {str(e)}")
                flash('处理开始日期时发生错误，请重新选择', 'error')
                return query
    
    # 处理结束日期
    if end_date:
        end_date = end_date.strip()
        
        if not end_date:
            # 空字符串，跳过处理
            pass
        elif not date_pattern.match(end_date):
            logging.warning(f"结束日期格式无效: {end_date}")
            flash('结束日期格式无效，请使用YYYY-MM-DD格式', 'error')
            return query
        else:
            try:
                end_dt = datetime.strptime(end_date, '%Y-%m-%d')
                
                # 验证日期合理性
                current_year = datetime.now().year
                if end_dt.year < 2020 or end_dt.year > current_year + 1:
                    logging.warning(f"结束日期超出合理范围: {end_date}")
                    flash('结束日期超出有效范围，请选择2020年到明年之间的日期', 'error')
                    return query
                
                # 设置为当天的最后一刻
                end_dt = end_dt.replace(hour=23, minute=59, second=59, microsecond=999999)
                query = query.filter(Order.created_at <= end_dt)
                logging.debug(f"应用结束日期筛选: {end_date}")
                
            except ValueError as e:
                logging.error(f"解析结束日期失败: {end_date}, 错误: {str(e)}")
                flash('结束日期无效，请检查日期是否存在（如2月30日）', 'error')
                return query
            except Exception as e:
                logging.error(f"处理结束日期时发生未知错误: {str(e)}")
                flash('处理结束日期时发生错误，请重新选择', 'error')
                return query
    
    # 验证日期范围逻辑
    if start_dt and end_dt:
        if start_dt.date() > end_dt.date():
            logging.info(f"用户输入了无效的日期范围: {start_date} > {end_date}")
            flash('开始日期不能大于结束日期，请重新选择', 'error')
            # 返回空结果集但保持查询结构
            return query.filter(Order.id == None)
        
        # 检查日期范围是否过大（超过2年）
        date_diff = end_dt.date() - start_dt.date()
        if date_diff.days > 730:  # 2年
            logging.warning(f"用户选择了过大的日期范围: {date_diff.days}天")
            flash('日期范围不能超过2年，请缩小查询范围', 'warning')
            # 不阻止查询，但给出警告
        
        logging.info(f"日期范围筛选: {start_date} 到 {end_date} ({date_diff.days}天)")
    
    return query

def apply_keyword_search(query, keyword):
    """应用关键词搜索 - 支持订单号、仓库、地址、货物、价格、供应商名称的模糊搜索"""
    if not keyword:
        return query
    
    conditions = [
        Order.order_no.ilike(f'%{keyword}%'),
        Order.warehouse.ilike(f'%{keyword}%'),
        Order.delivery_address.ilike(f'%{keyword}%'),
        Order.goods.ilike(f'%{keyword}%'),
        func.date(Order.created_at).like(f'%{keyword}%')
    ]
    
    # 价格搜索优化 - 精确匹配
    try:
        price_value = float(keyword)
        # 搜索中标价格（已完成订单）
        conditions.append(Order.selected_price == price_value)
        
        # 优化最低报价搜索 - 使用JOIN避免子查询性能问题
        Quote = Order._get_quote_model()
        price_match_orders = db.session.query(Order.id).join(Quote).filter(
            Quote.price == price_value
        ).subquery()
        conditions.append(Order.id.in_(price_match_orders))
        
    except (ValueError, TypeError):
        # 忽略非数字关键词的价格搜索
        pass
    
    # 供应商名称搜索 - 仅搜索已完成订单的中标供应商
    conditions.append(
        Order.selected_supplier.has(Supplier.name.ilike(f'%{keyword}%'))
    )
    
    return query.filter(or_(*conditions))

def notify_suppliers(order, suppliers):
    """通知供应商新订单 - 增强错误处理和重试机制"""
    success_count = 0
    failed_suppliers = []
    
    for supplier in suppliers:
        if not supplier.webhook_url:
            logging.info(f"供应商 {supplier.name} 未配置webhook，跳过通知")
            continue
            
        # 重试机制
        max_retries = 3
        for attempt in range(max_retries):
            try:
                access_url = url_for('supplier_portal', access_code=supplier.access_code, _external=True)
                
                # 验证访问码是否存在
                if not supplier.access_code:
                    logging.error(f"供应商 {supplier.name} 缺少访问码，无法生成链接")
                    failed_suppliers.append(supplier.name)
                    break
                
                message = {
                    "msgtype": "text",
                    "text": {
                        "content": f"🔔 新的询价订单通知\n\n"
                                   f"订单号：{order.order_no}\n"
                                   f"货物：{order.goods[:100]}...\n"  # 限制长度
                                   f"仓库：{order.warehouse}\n"
                                   f"收货地址：{order.delivery_address[:50]}...\n\n"
                                   f"请点击链接提交报价：{access_url}"
                    }
                }
                
                # 发送请求，设置超时
                response = requests.post(
                    supplier.webhook_url, 
                    json=message, 
                    timeout=5,  # 缩短超时时间
                    headers={'Content-Type': 'application/json'}
                )
                
                if response.status_code == 200:
                    logging.info(f"通知发送成功: {supplier.name} (尝试 {attempt + 1}/{max_retries})")
                    success_count += 1
                    break
                else:
                    logging.warning(f"通知发送失败: {supplier.name}, 状态码: {response.status_code}, 响应: {response.text[:200]}")
                    if attempt == max_retries - 1:
                        failed_suppliers.append(supplier.name)
                        
            except Timeout:
                logging.error(f"通知发送超时: {supplier.name} (尝试 {attempt + 1}/{max_retries})")
                if attempt == max_retries - 1:
                    failed_suppliers.append(supplier.name)
            except ConnectionError:
                logging.error(f"连接错误: {supplier.name} (尝试 {attempt + 1}/{max_retries})")
                if attempt == max_retries - 1:
                    failed_suppliers.append(supplier.name)
            except RequestException as e:
                logging.error(f"请求异常: {supplier.name}, 错误: {str(e)} (尝试 {attempt + 1}/{max_retries})")
                if attempt == max_retries - 1:
                    failed_suppliers.append(supplier.name)
            except Exception as e:
                logging.error(f"未知错误: {supplier.name}, 错误: {str(e)} (尝试 {attempt + 1}/{max_retries})")
                if attempt == max_retries - 1:
                    failed_suppliers.append(supplier.name)
            
            # 重试前等待
            if attempt < max_retries - 1:
                time.sleep(0.5 * (attempt + 1))  # 递增等待时间
    
    # 记录最终结果
    logging.info(f"供应商通知完成 - 成功: {success_count}, 失败: {len(failed_suppliers)}")
    if failed_suppliers:
        logging.error(f"通知失败的供应商: {', '.join(failed_suppliers)}")
    
    return success_count, failed_suppliers

@order_bp.route('/export')
@login_required
def export_orders():
    """导出订单列表为Excel文件 - 增强安全版本"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
        from flask import send_file
        import tempfile
        from utils.file_security import FileSecurity, file_security_check
        
        # 复用现有筛选逻辑
        status = request.args.get('status', '').strip()
        start_date = request.args.get('start_date', '').strip()
        end_date = request.args.get('end_date', '').strip()
        keyword = request.args.get('keyword', '').strip()
        
        # 构建查询 - 复用index()方法的筛选逻辑
        query = Order.query
        query = business_type_filter(query, Order)
        
        # 状态筛选
        if status and status in ['active', 'completed', 'cancelled']:
            query = query.filter_by(status=status)
        
        # 应用日期筛选
        query = apply_date_filter(query, start_date, end_date)
        
        # 应用关键词搜索
        query = apply_keyword_search(query, keyword)
        
        # 获取所有符合条件的订单
        orders = query.order_by(Order.created_at.desc()).all()
        
        if not orders:
            flash('没有符合条件的订单可以导出', 'warning')
            return redirect(url_for('order.index', status=status, start_date=start_date, end_date=end_date, keyword=keyword))
        
        # 检查数据量大小，如果超过1000条给出警告
        if len(orders) > 1000:
            logging.warning(f"用户{current_user.id}导出大量数据: {len(orders)}条记录")
        
        # 创建Excel工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "订单列表"
        
        # 设置标题样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 设置表头
        headers = ['订单号', '货物信息', '收货地址', '仓库', '报价数', '最低价/中标价', '供应商名称', '创建时间']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 填充数据
        for row, order in enumerate(orders, 2):
            ws.cell(row=row, column=1, value=order.order_no)
            ws.cell(row=row, column=2, value=order.goods)
            ws.cell(row=row, column=3, value=order.delivery_address)
            ws.cell(row=row, column=4, value=order.warehouse)
            ws.cell(row=row, column=5, value=order.get_quote_count())
            
            # 价格逻辑：已完成订单显示中标价，进行中订单显示最低价
            if order.status == 'completed' and order.selected_price:
                price_value = f"¥{order.selected_price:.2f}"
            elif order.status == 'active':
                lowest_quote = order.get_lowest_quote()
                price_value = f"¥{lowest_quote.price:.2f}" if lowest_quote else "-"
            else:
                price_value = "-"
            ws.cell(row=row, column=6, value=price_value)
            
            # 供应商名称：已完成订单显示中标供应商，进行中订单显示最低价供应商
            if order.status == 'completed' and order.selected_supplier:
                supplier_name = order.selected_supplier.name
            elif order.status == 'active':
                lowest_quote = order.get_lowest_quote()
                supplier_name = lowest_quote.supplier.name if lowest_quote and lowest_quote.supplier else "-"
            else:
                supplier_name = "-"
            ws.cell(row=row, column=7, value=supplier_name)
            
            ws.cell(row=row, column=8, value=order.created_at.strftime('%Y-%m-%d %H:%M'))
        
        # 自动调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 生成安全的文件名
        import datetime
        current_date = datetime.datetime.now().strftime('%y-%m-%d')
        raw_filename = f"订单{current_date}.xlsx"
        filename = FileSecurity.get_safe_filename(raw_filename)
        
        # 生成临时文件进行安全验证
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            wb.save(tmp_file.name)
            tmp_file_path = tmp_file.name
        
        # 文件安全验证
        is_valid, message = FileSecurity.validate_export_file(tmp_file_path)
        if not is_valid:
            os.unlink(tmp_file_path)  # 删除临时文件
            error_response, _ = ErrorHandler.handle_file_security_error(message, filename)
            ErrorResponseHelper.flash_error_message((error_response["error_code"], error_response["error_message"]))
            return redirect(url_for('order.index'))
        
        # 读取文件内容到内存
        with open(tmp_file_path, 'rb') as f:
            file_content = f.read()
        
        # 清理临时文件
        os.unlink(tmp_file_path)
        
        # 最终安全检查
        if len(file_content) > FileSecurity.MAX_FILE_SIZE:
            logging.error(f"导出文件过大: {len(file_content)}字节")
            ErrorResponseHelper.flash_error_message(ErrorCode.SEC_005, f"文件大小{len(file_content)}字节")
            return redirect(url_for('order.index'))
        
        # 创建内存文件对象
        excel_buffer = BytesIO(file_content)
        
        # 记录导出信息
        logging.info(f"Excel导出成功: 用户{current_user.id}, 导出{len(orders)}条记录, 文件大小:{len(file_content)}字节")
        
        # 返回文件
        return send_file(
            excel_buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logging.error(f"Excel导出失败: {str(e)}")
        logging.error(f"错误详情: {traceback.format_exc()}")
        flash('Excel导出失败，请稍后重试', 'error')
        return redirect(url_for('order.index'))