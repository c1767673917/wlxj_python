from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, session, send_file
from models import db, Supplier, Order, Quote
from datetime import datetime, date
from sqlalchemy import or_, func
from utils.beijing_time_helper import BeijingTimeHelper
# Excel导出相关导入
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
import logging
import traceback
from functools import wraps
import tempfile
import os
import psutil
from utils.file_security import FileSecurity, file_security_check

# 创建蓝图
portal_bp = Blueprint('portal', __name__, url_prefix='/portal')

# 供应商登录验证装饰器
def require_supplier_login(f):
    """供应商登录验证装饰器"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'supplier_id' not in session:
            flash('访问已过期，请重新访问', 'error')
            return redirect(url_for('index'))
        
        supplier_id = session['supplier_id']
        supplier = Supplier.query.get(supplier_id)
        
        if not supplier:
            flash('供应商信息不存在', 'error')
            return redirect(url_for('index'))
            
        return f(*args, **kwargs)
    return decorated_function

@portal_bp.route('/supplier/<access_code>')
def supplier_portal(access_code):
    """供应商专属门户入口"""
    supplier = Supplier.query.filter_by(access_code=access_code).first_or_404()
    
    # 将供应商信息存储到session中，用于后续验证
    session['supplier_id'] = supplier.id
    session['supplier_name'] = supplier.name
    session['access_code'] = access_code
    
    # 获取该供应商需要报价的订单（活跃状态且已关联的订单）
    orders = Order.query.join(Order.suppliers).filter(
        Supplier.id == supplier.id,
        Order.status == 'active'
    ).order_by(Order.created_at.desc()).all()
    
    # 获取该供应商已提交的报价
    quotes = Quote.query.filter_by(supplier_id=supplier.id).all()
    quoted_order_ids = [quote.order_id for quote in quotes]
    
    return render_template('portal/dashboard.html', 
                         supplier=supplier, 
                         orders=orders,
                         quoted_order_ids=quoted_order_ids,
                         quotes=quotes)

@portal_bp.route('/order/<int:order_id>')
@require_supplier_login
def order_detail(order_id):
    """订单详情页面（供应商视图）"""
    supplier_id = session['supplier_id']
    supplier = Supplier.query.get(supplier_id)
    
    # 获取订单并验证供应商是否有权限访问
    order = Order.query.join(Order.suppliers).filter(
        Order.id == order_id,
        Supplier.id == supplier_id
    ).first_or_404()
    
    # 获取该供应商对此订单的报价
    quote = Quote.query.filter_by(order_id=order.id, supplier_id=supplier_id).first()
    
    # 获取所有报价数量（不显示具体报价内容）
    total_quotes = Quote.query.filter_by(order_id=order.id).count()
    
    return render_template('portal/order_detail.html', 
                         supplier=supplier,
                         order=order, 
                         quote=quote,
                         total_quotes=total_quotes)

@portal_bp.route('/order/<int:order_id>/quote', methods=['GET', 'POST'])
@require_supplier_login
def submit_quote(order_id):
    """提交或更新报价"""
    supplier_id = session['supplier_id']
    supplier = Supplier.query.get(supplier_id)
    
    # 获取订单并验证权限
    order = Order.query.join(Order.suppliers).filter(
        Order.id == order_id,
        Supplier.id == supplier_id,
        Order.status == 'active'  # 只能对活跃订单报价
    ).first_or_404()
    
    # 获取现有报价
    existing_quote = Quote.query.filter_by(order_id=order.id, supplier_id=supplier_id).first()
    
    if request.method == 'POST':
        price = request.form.get('price', type=float)
        delivery_time = request.form.get('delivery_time', '').strip()
        remarks = request.form.get('remarks', '').strip()
        
        # 基础价格验证
        if not price or price <= 0:
            flash('请输入有效的报价金额', 'error')
            return render_template('portal/quote_form.html', 
                                 supplier=supplier, 
                                 order=order, 
                                 quote=existing_quote)
        
        # 价格上限验证
        if price > 9999999999.99:
            flash('报价金额超出系统允许的最大值', 'error')
            return render_template('portal/quote_form.html', 
                                 supplier=supplier, 
                                 order=order, 
                                 quote=existing_quote)
        
        # 价格变动合理性检查（仅在更新报价时）
        price_warnings = []
        if existing_quote:
            try:
                original_price = existing_quote.get_price_decimal()
                price_warnings = Quote.validate_price_change(original_price, price)
            except Exception as e:
                import logging
                logging.warning(f'价格对比计算失败: {e}')
        
        if existing_quote:
            # 更新现有报价
            existing_quote.price = price
            existing_quote.delivery_time = delivery_time if delivery_time else None
            existing_quote.remarks = remarks if remarks else None
            existing_quote.created_at = BeijingTimeHelper.now()  # 更新时间为北京时间
            
            # 显示价格警告和成功消息
            if price_warnings:
                for warning in price_warnings:
                    flash(warning, 'warning')
            flash('报价更新成功', 'success')
        else:
            # 创建新报价
            new_quote = Quote(
                order_id=order.id,
                supplier_id=supplier_id,
                price=price,
                delivery_time=delivery_time if delivery_time else None,
                remarks=remarks if remarks else None
            )
            db.session.add(new_quote)
            flash('报价提交成功', 'success')
        
        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            import logging
            logging.error(f'报价保存失败: {e}')
            flash('报价保存失败，请重试', 'error')
            return render_template('portal/quote_form.html', 
                                 supplier=supplier, 
                                 order=order, 
                                 quote=existing_quote)
        
        # 可以在这里添加通知采购方的逻辑
        notify_buyer_new_quote(order, supplier, price)
        
        return redirect(url_for('portal.order_detail', order_id=order.id))
    
    return render_template('portal/quote_form.html', 
                         supplier=supplier, 
                         order=order, 
                         quote=existing_quote)

@portal_bp.route('/quotes')
@require_supplier_login
def my_quotes():
    """我的报价列表 - 支持筛选、搜索和分页"""
    # 用于跟踪是否已经显示过用户提示，避免重复显示
    flash_message_shown = False
    
    try:
        supplier_id = session['supplier_id']
        supplier = Supplier.query.get(supplier_id)
        
        # 安全地获取筛选参数
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 10, type=int)
        status = request.args.get('status', '').strip()
        start_date = request.args.get('start_date', '').strip()
        end_date = request.args.get('end_date', '').strip()
        keyword = request.args.get('keyword', '').strip()
        date_quick = request.args.get('date_quick', '').strip()
        
        # 参数验证
        if page < 1:
            page = 1
        elif page > 1000:
            page = 1000
            
        if per_page not in [10, 20, 50]:
            per_page = 10
            
        # 验证状态参数
        valid_statuses = ['', 'active', 'completed', 'cancelled']
        if status not in valid_statuses:
            logging.warning(f"供应商提供了无效的状态参数: {status}")
            status = ''
        
        # 限制关键词长度
        if keyword and len(keyword) > 100:
            logging.warning(f"供应商提供了过长的关键词: {len(keyword)}字符")
            keyword = keyword[:100]
            flash('搜索关键词过长，已自动截取前100个字符', 'warning')
            flash_message_shown = True
        
        # 处理快捷日期选项
        if date_quick:
            quick_start, quick_end = process_quote_quick_date(date_quick)
            if quick_start and quick_end:
                start_date, end_date = quick_start, quick_end
                logging.debug(f"应用快捷日期选项: {date_quick} -> {start_date} to {end_date}")
            else:
                logging.warning(f"快捷日期选项处理失败: {date_quick}")
                flash('快捷日期设置失败，请手动选择日期', 'warning')
                flash_message_shown = True
        
        # 构建查询
        try:
            query = build_quotes_query(supplier_id, status, start_date, end_date, keyword)
        except ValueError as ve:
            # 用户输入验证错误，显示友好提示
            logging.warning(f"用户输入验证失败: {str(ve)}")
            flash(f'筛选条件有误：{str(ve)}', 'warning')
            flash_message_shown = True
            query = Quote.query.filter_by(supplier_id=supplier_id)
        except Exception as e:
            logging.error(f"构建报价查询失败: {str(e)}")
            flash('查询条件处理失败，显示所有报价', 'error')
            flash_message_shown = True
            query = Quote.query.filter_by(supplier_id=supplier_id)
        
        # 执行分页查询
        try:
            quotes = query.paginate(
                page=page, per_page=per_page, error_out=False
            )
            
            # 检查分页结果并记录信息
            if quotes.total == 0:
                if status or start_date or end_date or keyword:
                    logging.info(f"供应商{supplier_id}筛选条件未找到结果 - 状态:{status}, 日期:{start_date}-{end_date}, 关键词:{keyword}")
                    # 只有在之前没有显示过其他消息时才显示搜索结果提示
                    if not flash_message_shown:
                        flash('未找到符合条件的报价，请尝试调整搜索条件', 'info')
                        flash_message_shown = True
                else:
                    logging.info(f"供应商{supplier_id}暂无报价记录")
            else:
                logging.debug(f"供应商{supplier_id}查询完成，找到{quotes.total}条记录，显示第{page}页")
                
        except Exception as e:
            logging.error(f"分页查询执行失败: {str(e)}")
            logging.error(f"查询失败详情: {traceback.format_exc()}")
            
            # 尝试回退查询
            try:
                # 尝试最简单的查询：只获取供应商的报价，降低复杂度
                quotes = Quote.query.filter_by(supplier_id=supplier_id).order_by(
                    Quote.created_at.desc()).paginate(
                    page=1, per_page=10, error_out=False)
                
                if quotes.total == 0:
                    # 确实没有数据的情况
                    if not flash_message_shown:
                        flash('当前没有报价数据', 'info')
                        flash_message_shown = True
                else:
                    # 有数据但查询参数可能有问题
                    if not flash_message_shown:
                        flash('已为您显示最新的报价记录', 'info')
                        flash_message_shown = True
                    
                logging.info(f"供应商{supplier_id}使用简化查询成功，找到{quotes.total}条记录")
                
            except Exception as fallback_error:
                logging.error(f"回退查询也失败: {str(fallback_error)}")
                
                # 最后的兜底方案：创建空的分页对象
                quotes = create_mock_pagination_object(1, 10)
                if not flash_message_shown:
                    flash('系统暂时繁忙，请稍后再试', 'warning')
                    flash_message_shown = True
        
        return render_template('portal/quotes.html', 
                             supplier=supplier, 
                             quotes=quotes,
                             status=status,
                             start_date=start_date,
                             end_date=end_date,
                             keyword=keyword,
                             date_quick=date_quick,
                             per_page=per_page)
    
    except Exception as e:
        logging.error(f"供应商报价列表页面发生未知错误: {str(e)}")
        logging.error(f"错误详情: {traceback.format_exc()}")
        
        # 只有在之前没有显示过其他消息时才显示紧急模式提示
        try:
            supplier_id = session.get('supplier_id')
            supplier = Supplier.query.get(supplier_id) if supplier_id else None
            
            if supplier:
                # 创建最基础的页面显示
                quotes = create_mock_pagination_object(1, 10)
                # 只有在没有其他flash消息时才显示紧急模式提示
                if not flash_message_shown:
                    flash('页面加载遇到问题，请稍后再试', 'warning')
                
                return render_template('portal/quotes.html', 
                                     supplier=supplier, 
                                     quotes=quotes,
                                     status='',
                                     start_date='',
                                     end_date='',
                                     keyword='',
                                     date_quick='',
                                     per_page=10)
        except Exception as emergency_error:
            logging.error(f"紧急模式也失败: {str(emergency_error)}")
        
        # 最终兜底：跳转到供应商门户
        flash('页面加载失败，请重新进入系统', 'error')
        try:
            access_code = session.get('access_code', '')
            if access_code:
                return redirect(url_for('portal.supplier_portal', access_code=access_code))
            else:
                # 如果没有access_code，清理session并跳转到首页
                session.clear()
                flash('会话已过期，请重新访问', 'info')
                return redirect(url_for('index'))
        except Exception as redirect_error:
            logging.error(f"错误页面跳转失败: {str(redirect_error)}")
            session.clear()
            return redirect(url_for('index'))

@portal_bp.route('/logout')
def logout():
    """供应商登出"""
    session.pop('supplier_id', None)
    session.pop('supplier_name', None)
    session.pop('access_code', None)
    flash('您已安全退出', 'info')
    return redirect(url_for('index'))

# ====== 供应商报价Excel导出相关函数 ======

def prepare_quotes_export_data(supplier_id, status, start_date, end_date, keyword):
    """准备供应商报价导出数据"""
    try:
        supplier = Supplier.query.get(supplier_id)
        if not supplier:
            return None, None, "供应商信息无效"
        
        logging.info(f"供应商 {supplier_id} 开始报价导出准备")
        logging.debug(f"供应商导出筛选条件: status={status}, start_date={start_date}, end_date={end_date}, keyword={keyword}")
        
        # 构建查询 - 复用筛选逻辑
        query = build_quotes_query(supplier_id, status, start_date, end_date, keyword)
        
        # 获取总数量进行评估
        total_count = query.count()
        logging.info(f"供应商 {supplier_id} 查询到 {total_count} 个符合条件的报价")
        
        if total_count == 0:
            return None, supplier, "没有符合条件的报价可以导出"
        
        # 内存监控
        try:
            process = psutil.Process()
            memory_info = process.memory_info()
            logging.debug(f"当前内存使用: {memory_info.rss/1024/1024:.2f}MB")
        except Exception as e:
            logging.warning(f"内存监控失败: {e}")
        
        return query, supplier, None
        
    except Exception as e:
        logging.error(f"供应商报价数据准备失败: {str(e)}")
        return None, None, f"数据准备失败: {str(e)}"

def create_quotes_excel_workbook(supplier):
    """创建供应商报价Excel工作簿"""
    try:
        logging.debug("开始创建供应商报价Excel工作簿")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{supplier.name}报价列表"
        
        # 设置标题样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 设置表头
        headers = ['订单号', '货物信息', '收货地址', '仓库', '报价金额', '交期', '订单状态', '报价状态', '创建时间']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        logging.debug("供应商报价表头设置完成")
        return wb, ws, headers
        
    except Exception as e:
        logging.error(f"创建供应商报价Excel工作簿失败: {str(e)}")
        raise

def fill_quotes_excel_data(ws, query, supplier_id, batch_size=200):
    """填充供应商报价数据 - 分批处理"""
    try:
        logging.debug("开始填充供应商报价数据")
        
        total_processed = 0
        current_row = 2  # 从第二行开始
        
        # 分批处理数据
        offset = 0
        while True:
            # 获取一批数据
            batch_quotes = query.offset(offset).limit(batch_size).all()
            
            if not batch_quotes:
                break
                
            logging.debug(f"处理供应商报价第 {offset//batch_size + 1} 批，{len(batch_quotes)} 条记录")
            
            # 处理当前批次的数据
            for quote in batch_quotes:
                try:
                    # 基本信息
                    ws.cell(row=current_row, column=1, value=quote.order.order_no if quote.order else "-")
                    ws.cell(row=current_row, column=2, value=quote.order.goods if quote.order else "-")
                    ws.cell(row=current_row, column=3, value=quote.order.delivery_address if quote.order else "-")
                    ws.cell(row=current_row, column=4, value=quote.order.warehouse if quote.order else "-")
                    
                    # 报价信息
                    try:
                        price_str = f"￥{quote.price:.2f}" if quote.price else "-"
                        ws.cell(row=current_row, column=5, value=price_str)
                    except Exception as e:
                        logging.warning(f"格式化报价金额失败: {e}")
                        ws.cell(row=current_row, column=5, value="-")
                    
                    ws.cell(row=current_row, column=6, value=quote.delivery_time or '-')
                    
                    # 订单状态
                    try:
                        status_map = {
                            'active': '进行中',
                            'completed': '已完成', 
                            'cancelled': '已取消'
                        }
                        order_status = status_map.get(quote.order.status, quote.order.status) if quote.order else "-"
                        ws.cell(row=current_row, column=7, value=order_status)
                    except Exception as e:
                        logging.warning(f"获取订单状态失败: {e}")
                        ws.cell(row=current_row, column=7, value="-")
                    
                    # 报价状态
                    try:
                        if quote.order and quote.order.selected_supplier_id == supplier_id:
                            quote_status = '已中标'
                        elif quote.order and quote.order.status == 'completed':
                            quote_status = '未中标'
                        elif quote.order and quote.order.status == 'cancelled':
                            quote_status = '订单取消'
                        else:
                            quote_status = '待定'
                        ws.cell(row=current_row, column=8, value=quote_status)
                    except Exception as e:
                        logging.warning(f"获取报价状态失败: {e}")
                        ws.cell(row=current_row, column=8, value="-")
                    
                    # 创建时间
                    try:
                        created_time = quote.created_at.strftime('%Y-%m-%d %H:%M') if quote.created_at else "-"
                        ws.cell(row=current_row, column=9, value=created_time)
                    except Exception as e:
                        logging.warning(f"格式化创建时间失败: {e}")
                        ws.cell(row=current_row, column=9, value="-")
                    
                    current_row += 1
                    total_processed += 1
                    
                except Exception as e:
                    logging.error(f"处理报价数据时出错: {e}")
                    # 继续处理下一个报价，而不是中断整个导出
                    continue
            
            offset += batch_size
        
        logging.debug(f"供应商报价数据填充完成，处理了 {total_processed} 条记录")
        return total_processed
        
    except Exception as e:
        logging.error(f"供应商报价数据填充失败: {str(e)}")
        raise

def finalize_quotes_export(wb, supplier, total_records):
    """完成供应商报价文件生成和验证"""
    temp_file_path = None
    try:
        # 生成安全的文件名
        current_date = BeijingTimeHelper.get_backup_timestamp()
        raw_filename = f"{supplier.name}报价导出_{current_date}.xlsx"
        filename = FileSecurity.get_safe_filename(raw_filename)
        
        logging.debug(f"生成供应商报价文件名: {filename}")
        
        # 保存到临时文件
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            temp_file_path = tmp_file.name
            wb.save(temp_file_path)
        
        logging.debug(f"供应商报价Excel文件保存到临时路径: {temp_file_path}")
        
        # 文件安全验证
        is_valid, message = FileSecurity.validate_export_file(temp_file_path)
        if not is_valid:
            logging.error(f"供应商报价文件安全验证失败: {message}")
            if temp_file_path and os.path.exists(temp_file_path):
                os.unlink(temp_file_path)
            return None, None, f"文件安全验证失败: {message}"
        
        # 读取文件内容到内存
        try:
            with open(temp_file_path, 'rb') as f:
                file_content = f.read()
            logging.debug(f"供应商报价文件内容读取成功，大小: {len(file_content)} bytes")
        except Exception as e:
            logging.error(f"读取供应商报价临时文件失败: {e}")
            if temp_file_path and os.path.exists(temp_file_path):
                os.unlink(temp_file_path)
            return None, None, f"文件读取失败: {str(e)}"
        
        # 清理临时文件
        if temp_file_path and os.path.exists(temp_file_path):
            os.unlink(temp_file_path)
            temp_file_path = None
        
        # 最终安全检查
        if len(file_content) > FileSecurity.MAX_FILE_SIZE:
            logging.error(f"供应商报价导出文件过大: {len(file_content)}字节")
            return None, None, f"导出文件过大: {len(file_content)}字节"
        
        # 创建内存文件对象
        excel_buffer = BytesIO(file_content)
        
        # 记录导出信息
        logging.info(f"供应商{supplier.id}报价Excel导出成功: 导出{total_records}条记录, 文件大小:{len(file_content)}字节")
        
        return excel_buffer, filename, None
        
    except Exception as e:
        # 确保临时文件被清理
        if temp_file_path and os.path.exists(temp_file_path):
            try:
                os.unlink(temp_file_path)
            except Exception:
                pass
        
        logging.error(f"供应商报价文件生成失败: {str(e)}")
        return None, None, f"文件生成失败: {str(e)}"

@portal_bp.route('/quotes/export')
@require_supplier_login
@file_security_check
def export_quotes():
    """导出筛选后的报价Excel - 重构版本支持分批处理和性能优化"""
    try:
        supplier_id = session['supplier_id']
        
        # 获取筛选参数
        status = request.args.get('status', '').strip()
        start_date = request.args.get('start_date', '').strip()
        end_date = request.args.get('end_date', '').strip()
        keyword = request.args.get('keyword', '').strip()
        
        # 步骤1: 数据准备和查询
        query, supplier, error_msg = prepare_quotes_export_data(supplier_id, status, start_date, end_date, keyword)
        if error_msg:
            flash(error_msg, 'warning')
            return redirect(url_for('portal.my_quotes', 
                                  status=status, start_date=start_date, 
                                  end_date=end_date, keyword=keyword))
        
        # 步骤2: 创建Excel工作簿
        wb, ws, headers = create_quotes_excel_workbook(supplier)
        
        # 步骤3: 数据填充（分批处理）
        total_records = fill_quotes_excel_data(ws, query, supplier_id, batch_size=200)
        
        if total_records == 0:
            flash('没有数据可导出', 'warning')
            return redirect(url_for('portal.my_quotes', 
                                  status=status, start_date=start_date, 
                                  end_date=end_date, keyword=keyword))
        
        # 步骤4: 优化Excel格式
        optimize_quotes_excel_formatting(ws)
        
        # 步骤5: 文件生成和验证
        excel_buffer, filename, error_msg = finalize_quotes_export(wb, supplier, total_records)
        if error_msg:
            flash(f'Excel导出失败: {error_msg}', 'error')
            return redirect(url_for('portal.my_quotes'))
        
        # 返回文件
        return send_file(
            excel_buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except ImportError as e:
        logging.error(f"模块导入失败: {str(e)}")
        logging.error(f"错误详情: {traceback.format_exc()}")
        flash('系统组件加载失败，请联系管理员', 'error')
        return redirect(url_for('portal.my_quotes'))
    except Exception as e:
        logging.error(f"供应商报价Excel导出失败: {str(e)}")
        logging.error(f"错误详情: {traceback.format_exc()}")
        flash('Excel导出失败，请稍后重试', 'error')
        return redirect(url_for('portal.my_quotes'))

def optimize_quotes_excel_formatting(ws):
    """优化供应商报价Excel格式"""
    try:
        logging.debug("开始优化供应商报价Excel格式")
        
        # 自动调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    cell_value = str(cell.value) if cell.value is not None else ""
                    if len(cell_value) > max_length:
                        max_length = len(cell_value)
                except Exception:
                    continue
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        logging.debug("供应商报价列宽调整完成")
        
    except Exception as e:
        logging.warning(f"调整供应商报价列宽时出错: {e}")
        # 列宽调整失败不影响导出

def build_quotes_query(supplier_id, status=None, start_date=None, end_date=None, keyword=None):
    """构建报价查询条件"""
    # 基础查询 - 连接Order表进行筛选
    query = Quote.query.join(Order).filter(Quote.supplier_id == supplier_id)
    
    # 应用筛选条件
    query = apply_quote_filters(query, status, start_date, end_date, keyword)
    
    return query.order_by(Quote.created_at.desc())

def apply_quote_filters(query, status, start_date, end_date, keyword):
    """应用报价筛选条件"""
    # 状态筛选
    if status:
        query = query.filter(Order.status == status)
        logging.debug(f"应用状态筛选: {status}")
    
    # 日期范围筛选
    try:
        query = apply_quote_date_filter(query, start_date, end_date)
    except Exception as e:
        logging.error(f"日期筛选处理失败: {str(e)}")
        raise e
    
    # 关键词搜索
    try:
        query = apply_quote_keyword_search(query, keyword)
    except Exception as e:
        logging.error(f"关键词搜索处理失败: {str(e)}")
        raise e
    
    return query

def apply_quote_date_filter(query, start_date, end_date):
    """应用报价日期范围筛选 - 增强版本，提供更好的错误处理"""
    import re
    from datetime import timedelta
    
    start_dt = None
    end_dt = None
    
    # 日期格式验证
    date_pattern = re.compile(r'^\d{4}-\d{2}-\d{2}$')
    
    # 处理开始日期
    if start_date and start_date.strip():
        start_date = start_date.strip()
        if not date_pattern.match(start_date):
            raise ValueError(f'开始日期格式无效，请使用 YYYY-MM-DD 格式')
        
        try:
            start_dt = datetime.strptime(start_date, '%Y-%m-%d')
            # 验证日期的合理性
            if start_dt.year < 2000 or start_dt.year > 2100:
                raise ValueError('开始日期年份超出合理范围（2000-2100）')
            query = query.filter(Order.created_at >= start_dt)
            logging.debug(f"应用开始日期筛选: {start_date}")
        except ValueError as e:
            if "格式" in str(e) or "范围" in str(e):
                raise e
            else:
                raise ValueError(f'开始日期无效：{start_date}，请检查日期是否存在')
    
    # 处理结束日期
    if end_date and end_date.strip():
        end_date = end_date.strip()
        if not date_pattern.match(end_date):
            raise ValueError(f'结束日期格式无效，请使用 YYYY-MM-DD 格式')
        
        try:
            end_dt = datetime.strptime(end_date, '%Y-%m-%d')
            # 验证日期的合理性
            if end_dt.year < 2000 or end_dt.year > 2100:
                raise ValueError('结束日期年份超出合理范围（2000-2100）')
            # 设置为当天的最后一刻
            end_dt = end_dt.replace(hour=23, minute=59, second=59, microsecond=999999)
            query = query.filter(Order.created_at <= end_dt)
            logging.debug(f"应用结束日期筛选: {end_date}")
        except ValueError as e:
            if "格式" in str(e) or "范围" in str(e):
                raise e
            else:
                raise ValueError(f'结束日期无效：{end_date}，请检查日期是否存在')
    
    # 验证日期范围
    if start_dt and end_dt:
        if start_dt.date() > end_dt.date():
            raise ValueError('开始日期不能大于结束日期')
        
        date_diff = end_dt.date() - start_dt.date()
        if date_diff.days > 365:  # 1年
            logging.warning(f"供应商选择了过大的日期范围: {date_diff.days}天")
            raise ValueError('选择的日期范围不能超过1年，请缩小查询范围')
    
    return query

def apply_quote_keyword_search(query, keyword):
    """应用关键词搜索 - 搜索订单号、仓库、收货地址、货物信息，增强版本"""
    if not keyword or not keyword.strip():
        return query
    
    keyword = keyword.strip()
    
    # 关键词长度验证
    if len(keyword) > 100:
        raise ValueError('搜索关键词长度不能超过100个字符')
    
    # 关键词内容验证 - 防止SQL注入等
    import re
    # 允许中文、英文、数字、常见符号
    if not re.match(r'^[\w\u4e00-\u9fff\s\-_.,()（）【】\[\]]+$', keyword):
        raise ValueError('搜索关键词包含不支持的字符，请使用中文、英文、数字或常见符号')
    
    # 过滤掉过短的关键词
    if len(keyword) < 1:
        return query
    
    # 如果关键词过短，给出提示但仍然搜索
    if len(keyword) == 1:
        logging.info(f"供应商使用了单字符搜索关键词: '{keyword}'")
    
    try:
        search_pattern = f"%{keyword}%"
        
        # 多字段模糊匹配
        conditions = [
            Order.order_no.ilike(search_pattern),
            Order.goods.ilike(search_pattern), 
            Order.delivery_address.ilike(search_pattern),
            Order.warehouse.ilike(search_pattern)
        ]
        
        return query.filter(or_(*conditions))
        
    except Exception as e:
        logging.error(f"关键词搜索应用失败: {e}")
        raise ValueError('搜索关键词处理失败，请重新输入')

def process_quote_quick_date(date_quick):
    """处理供应商报价快捷日期选项"""
    try:
        today = date.today()
        
        if date_quick == 'today':
            date_str = today.strftime('%Y-%m-%d')
            return date_str, date_str
        elif date_quick == 'this_month':
            start = today.replace(day=1)
            start_str = start.strftime('%Y-%m-%d')
            end_str = today.strftime('%Y-%m-%d')
            return start_str, end_str
        else:
            logging.warning(f"不支持的快捷日期选项: {date_quick}")
            return '', ''
            
    except Exception as e:
        logging.error(f"处理快捷日期选项时发生错误: {str(e)}")
        return '', ''

def notify_buyer_new_quote(order, supplier, price):
    """通知采购方有新报价（可扩展实现）"""
    # 这里可以实现通知采购方的逻辑
    # 例如：发送邮件、系统内通知等
    pass

def create_mock_pagination_object(page=1, per_page=10):
    """创建完全兼容Flask-SQLAlchemy Pagination接口的Mock对象
    
    确保与模板中使用的所有Pagination属性和方法完全兼容
    """
    
    class MockPagination:
        """完全兼容的Pagination模拟对象"""
        
        def __init__(self, page=1, per_page=10):
            # 基础分页属性
            self.items = []
            self.total = 0
            self.pages = 0
            self.page = page
            self.per_page = per_page
            
            # 分页导航属性
            self.has_prev = False
            self.has_next = False
            self.prev_num = None
            self.next_num = None
            
            # 附加属性（确保完全兼容）
            self.query = None
            self.error_out = False
            
        def iter_pages(self, left_edge=2, left_current=2, right_current=3, right_edge=2):
            """模拟分页页码迭代器，与Flask-SQLAlchemy完全兼容"""
            return []
            
        def __iter__(self):
            """支持直接迭代分页对象"""
            return iter(self.items)
            
        def __len__(self):
            """支持len()函数"""
            return len(self.items)
            
        def __bool__(self):
            """支持布尔值判断"""
            return len(self.items) > 0
            
        def __nonzero__(self):
            """Python 2兼容性"""
            return self.__bool__()
            
        def __repr__(self):
            """字符串表示"""
            return f'<MockPagination page={self.page} total={self.total}>'
            
        # 兼容性方法：确保与真实Pagination对象的所有接口一致
        @property 
        def prev(self):
            """前一页对象（为空）"""
            return None
            
        @property
        def next(self):
            """下一页对象（为空）"""
            return None
    
    try:
        # 验证参数合法性
        if page < 1:
            page = 1
        if per_page < 1:
            per_page = 10
            
        mock_obj = MockPagination(page, per_page)
        
        logging.debug(f"创建MockPagination对象成功: page={page}, per_page={per_page}")
        return mock_obj
        
    except Exception as e:
        logging.error(f"创建MockPagination对象失败: {str(e)}")
        # 返回最简单的备用对象
        return type('FallbackPagination', (), {
            'items': [], 'total': 0, 'pages': 0, 
            'has_prev': False, 'has_next': False,
            'prev_num': None, 'next_num': None,
            'page': 1, 'per_page': 10,
            'iter_pages': lambda: [],
            '__iter__': lambda self: iter([]),
            '__len__': lambda self: 0,
            '__bool__': lambda self: False
        })()